import pandas as pd
import openpyxl
from datetime import datetime
import os

def colar_valor_ao_lado(ws, texto_alvo, valor, ocorrencia=1):
    cont = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and texto_alvo.lower() in cell.value.lower():
                cont += 1
                if cont == ocorrencia:
                    col_destino = cell.column + 1
                    linha_destino = cell.row
                    ws.cell(row=linha_destino, column=col_destino).value = valor
                    return True
    return False

def processar_planilhas():
    caminho_origem = r"C:\Users\wendel.ferreira\Desktop\TESTAT\RelatóriodeAudiências25.xlsx"
    caminho_destino = r"C:\Users\wendel.ferreira\Desktop\TESTAT\TRT04.08a08.08.xlsx"

    try:
        if not os.path.exists(caminho_origem) or not os.path.exists(caminho_destino):
            print("Erro: Um dos arquivos não foi encontrado.")
            return

        df_origem = pd.read_excel(caminho_origem, sheet_name='Sheet1')

        mapeamento_colunas = {
            'DATA': 'DATA',
            'HORARIO': 'HORARIO', 
            'RECLAMANTE': 'RECLAMANTE',
            'RECLAMADO': 'RECLAMADO',
            'Nº DO PROCESSO': 'NUM_PROCESSO',
            'TURMA': 'TURMA',
            'RELATOR': 'RELATOR',
            'TIPO DE RECURSO': 'TIPO_RECURSO'
        }

        df_origem = df_origem.rename(columns=mapeamento_colunas)
        colunas_necessarias = ['DATA', 'HORARIO', 'RECLAMANTE', 'RECLAMADO', 'NUM_PROCESSO', 'TURMA', 'RELATOR', 'TIPO_RECURSO']
        df_filtrado = df_origem[colunas_necessarias].copy()
        df_filtrado = df_filtrado.dropna(subset=['DATA', 'HORARIO'])

        # Conversão e ordenação
        df_filtrado['DATA'] = pd.to_datetime(df_filtrado['DATA'], errors='coerce')
        df_filtrado['HORARIO_TIME'] = pd.to_datetime(df_filtrado['HORARIO'], format='%H:%M', errors='coerce').dt.time
        df_filtrado = df_filtrado.sort_values(['DATA', 'HORARIO_TIME'])

        wb_destino = openpyxl.load_workbook(caminho_destino)
        ws_destino = wb_destino["Julho"] if "Julho" in wb_destino.sheetnames else wb_destino.active

        ocorrencia_global = 1

        # Agrupar por data para manter ordem por dia
        for data, grupo in df_filtrado.groupby('DATA'):
            grupo_ordenado = grupo.sort_values(by='HORARIO')

            for _, row in grupo_ordenado.iterrows():
                tipo = str(row.get('TIPO_RECURSO', '')).strip().upper()
                if "RECURSO ORDINARIO" in tipo:
                    prefixo = "RO "
                elif "AGRAVO DE PETIÇÃO" in tipo:
                    prefixo = "AP "
                else:
                    prefixo = ""

                numero_formatado = f"{prefixo}{str(row['NUM_PROCESSO'])}"

                colar_valor_ao_lado(ws_destino, "Nº processo:", numero_formatado, ocorrencia=ocorrencia_global)
                colar_valor_ao_lado(ws_destino, "Reclamante:", str(row['RECLAMANTE']), ocorrencia=ocorrencia_global)
                colar_valor_ao_lado(ws_destino, "Horário:", str(row['HORARIO']), ocorrencia=ocorrencia_global)
                colar_valor_ao_lado(ws_destino, "Relator:", str(row['RELATOR']), ocorrencia=ocorrencia_global)
                colar_valor_ao_lado(ws_destino, "Reclamado:", str(row['RECLAMADO']), ocorrencia=ocorrencia_global)

                turma_texto = str(row['TURMA']).strip()
                partes = turma_texto.split()
                numero_turma = partes[0] if len(partes) >= 1 else turma_texto
                local_turma = " ".join(partes[1:]) if len(partes) > 1 else ""

                colar_valor_ao_lado(ws_destino, "Turma:", numero_turma, ocorrencia=ocorrencia_global)
                colar_valor_ao_lado(ws_destino, "Local:", local_turma, ocorrencia=ocorrencia_global)

                ocorrencia_global += 1  # Incrementa depois de cada registro

        wb_destino.save(caminho_destino)
        wb_destino.close()
        print("Automação concluída com sucesso.")

    except Exception as e:
        print(f"Erro durante a execução: {e}")

if __name__ == "__main__":
    processar_planilhas()
